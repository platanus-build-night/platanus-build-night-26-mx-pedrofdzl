import csv
import io

from openpyxl import load_workbook

from core.models import Requirement
from core.services.embeddings import embed_texts
from core.services.questionnaire_mapper import map_layout


def normalized_key(text):
    return " ".join(text.lower().split())


def file_format(filename):
    return "csv" if filename.lower().endswith(".csv") else "xlsx"


def read_rows(file_obj, filename):
    if file_format(filename) == "csv":
        return _csv_rows(file_obj)
    return _xlsx_rows(file_obj)


def _csv_rows(file_obj):
    raw = file_obj.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8")
    return [list(row) for row in csv.reader(io.StringIO(raw))]


def _xlsx_rows(file_obj):
    sheet = load_workbook(file_obj, read_only=True, data_only=True).active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def _cell(row, idx):
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    value = row[idx]
    return "" if value is None else str(value).strip()


def ingest_questionnaire(questionnaire, file_obj=None, filename=None):
    file_obj = file_obj if file_obj is not None else questionnaire.raw_file
    filename = filename or getattr(file_obj, "name", "")
    if hasattr(file_obj, "open"):
        file_obj.open("rb")

    rows = read_rows(file_obj, filename)
    layout = map_layout(rows)
    layout["format"] = file_format(filename)

    header_row = layout.get("header_row")
    question_col = layout.get("question_column", 0)
    category_col = layout.get("category_column")
    start = header_row + 1 if isinstance(header_row, int) and header_row >= 0 else 0

    requirements = []
    for index in range(start, len(rows)):
        text = _cell(rows[index], question_col)
        if not text:
            continue
        requirements.append(
            Requirement(
                questionnaire=questionnaire,
                text=text,
                category=_cell(rows[index], category_col),
                normalized_key=normalized_key(text),
                source_row=index,
            )
        )
    created = Requirement.objects.bulk_create(requirements)
    if created:
        for req, vector in zip(created, embed_texts([r.text for r in created]), strict=False):
            req.embedding = vector
        Requirement.objects.bulk_update(created, ["embedding"])
    questionnaire.layout = layout
    questionnaire.save(update_fields=["layout"])
    return {"requirements": len(created), "layout": layout}
