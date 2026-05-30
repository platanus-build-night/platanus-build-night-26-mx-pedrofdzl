import csv
import io

from openpyxl import load_workbook

from core.models import Answer

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_CONTENT_TYPE = "text/csv"


def _best_answer(requirement):
    answers = list(requirement.answers.all())
    if not answers:
        return None
    approved = [a for a in answers if a.status == Answer.Status.APPROVED]
    return max(approved or answers, key=lambda a: a.created_at)


def _format(questionnaire):
    fmt = (questionnaire.layout or {}).get("format")
    if fmt:
        return fmt
    name = questionnaire.raw_file.name or ""
    return "csv" if name.lower().endswith(".csv") else "xlsx"


def _answers_by_row(questionnaire):
    requirements = questionnaire.requirements.prefetch_related("answers")
    result = {}
    for req in requirements:
        if req.source_row is None:
            continue
        answer = _best_answer(req)
        if answer is not None:
            result[req.source_row] = answer.text
    return result


def export_questionnaire(questionnaire):
    if _format(questionnaire) == "csv":
        return _export_csv(questionnaire)
    return _export_xlsx(questionnaire)


def _export_xlsx(questionnaire):
    layout = questionnaire.layout or {}
    answers = _answers_by_row(questionnaire)
    questionnaire.raw_file.open("rb")
    workbook = load_workbook(questionnaire.raw_file)
    sheet = workbook.active

    answer_col = layout.get("answer_column")
    if answer_col is None:
        answer_col = sheet.max_column
        header_row = layout.get("header_row")
        if isinstance(header_row, int) and header_row >= 0:
            sheet.cell(row=header_row + 1, column=answer_col + 1, value="Answer")

    for source_row, text in answers.items():
        sheet.cell(row=source_row + 1, column=answer_col + 1, value=text)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), XLSX_CONTENT_TYPE, f"{questionnaire.source_name}-filled.xlsx"


def _set_cell(rows, r, c, value):
    while len(rows[r]) <= c:
        rows[r].append("")
    rows[r][c] = value


def _export_csv(questionnaire):
    layout = questionnaire.layout or {}
    answers = _answers_by_row(questionnaire)
    questionnaire.raw_file.open("rb")
    raw = questionnaire.raw_file.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8")
    rows = [list(row) for row in csv.reader(io.StringIO(raw))]

    answer_col = layout.get("answer_column")
    if answer_col is None:
        answer_col = max((len(r) for r in rows), default=0)
        header_row = layout.get("header_row")
        if isinstance(header_row, int) and 0 <= header_row < len(rows):
            _set_cell(rows, header_row, answer_col, "Answer")

    for source_row, text in answers.items():
        if 0 <= source_row < len(rows):
            _set_cell(rows, source_row, answer_col, text)

    buffer = io.StringIO()
    csv.writer(buffer).writerows(rows)
    filename = f"{questionnaire.source_name}-filled.csv"
    return buffer.getvalue().encode("utf-8"), CSV_CONTENT_TYPE, filename
